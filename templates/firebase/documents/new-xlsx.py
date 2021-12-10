import urllib.request
import json
from openpyxl.styles import colors
from openpyxl.styles import Font
from openpyxl import Workbook, load_workbook

word_levels = ('w1', 'w2', 'w3')
column_names = ['級別', '單字', '意義']
url = 'http://localhost:5000/gept/gept-words_json/{}'
filename = 'gept-words-{}.xlsx'

for wl in word_levels:
    req = urllib.request.urlopen(url.format(wl))
    response_data = req.read()
    json_data = json.loads(response_data)

    wb = Workbook()
    ws = wb.active

    # 設定欄位
    ws.append(column_names)

    for letter in json_data['jsonData']:
        for word in letter['words']:
            ws.append(word)
    
    # 設定字型
    font_face = 'Consolas'
    for i, row in enumerate(ws.rows):
        if i == 0:
            for cell in row:
                cell.font = Font(name=font_face, size=12, color=colors.BLUE)
        else:
            for cell in row:
                cell.font = Font(name=font_face, size=12)

    wb.save(filename.format(wl))

    # 再次讀取
    wb = load_workbook(filename.format(wl))
    ws = wb.active
    for ri, row in enumerate(list(ws.rows)[:4]):
        if ri > 0:
            for ci, cell in enumerate(row):
                if ci > 0:
                    print(cell.value)
    print()
    wb.save(filename.format(wl))
